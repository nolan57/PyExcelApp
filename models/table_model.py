import threading
from typing import Any, Optional, List
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex, pyqtSlot, QMetaObject
from PyQt6.QtGui import QColor, QBrush
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from utils.error_handler import ErrorHandler
from globals import GlobalState
from plugin_manager.features.plugin_permissions import PluginPermission
import logging

class TableModel(QAbstractTableModel):
    """Excel工作表的数据模型"""
    
    def __init__(self, worksheet: Worksheet):
        super().__init__()
        self.worksheet = worksheet
        self._data = {}  # 缓存修改的数据
        self._colors = {}  # 单元格颜色
        self._original_data = {}  # 保存原始数据用于比较
        self._max_row = worksheet.max_row
        self._max_column = worksheet.max_column
        self._save_lock = threading.Lock()  # 用于保护保存操作的线程锁
        self.load_data()

    def load_data(self):
        """从worksheet加载数据到缓存"""
        try:
            # 清除现有数据
            self._data.clear()
            self._colors.clear()
            self._original_data.clear()
            
            # 使用 worksheet.iter_rows() 替代直接访问，提高大文件加载性能
            for row in self.worksheet.iter_rows(min_row=1, max_row=self._max_row, 
                                              max_col=self._max_column):
                for cell in row:
                    if cell.value is not None:  # 只缓存非空单元格
                        self._data[(cell.row-1, cell.column-1)] = str(cell.value)
                        
                        # 获取单元格颜色
                        # if cell.fill and cell.fill.start_color:
                        #     color = cell.fill.start_color.rgb
                        #     if color:
                        #         self._colors[(cell.row-1, cell.column-1)] = QColor(color)
                                
            # 保存原始数据的副本
            self._original_data = self._data.copy()
                
        except Exception as e:
            ErrorHandler.handle_error(e, None, "加载表格数据时发生错误")
            
    def rowCount(self, parent=QModelIndex()) -> int:
        """返回行数"""
        return self._max_row
        
    def columnCount(self, parent=QModelIndex()) -> int:
        """返回列数"""
        return self._max_column
        
    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole) -> Any:
        """获取单元格数据"""
        if not index.isValid():
            return None
            
        row, col = index.row(), index.column()
        
        if role == Qt.ItemDataRole.DisplayRole or role == Qt.ItemDataRole.EditRole:
            # 优先返回修改的数据
            return self._data.get((row, col), self.get_cell_value(row, col))
            
        elif role == Qt.ItemDataRole.BackgroundRole:
            # 返回单元格背景色
            return QBrush(self._colors.get((row, col), Qt.GlobalColor.white))
            
        return None
    
    def setData(self, index: QModelIndex, value: Any, role: int = Qt.ItemDataRole.EditRole) -> bool:
        """设置单元格数据"""
        if not index.isValid() or role != Qt.ItemDataRole.EditRole:
            return False
            
        row, col = index.row(), index.column()
        
        # 更新缓存的数据
        if value:
            self._logger.info(f"设置单元格数据: {row}, {col}, {value}")
            self._data[(row, col)] = str(value)
        else:
            self._data.pop((row, col), None)
            
        # 发出数据更改信号
        self.dataChanged.emit(index, index, [role])
        return True
        
    def flags(self, index: QModelIndex) -> Qt.ItemFlag:
        """返回单元格标志"""
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags
            
        return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEditable
        
    def get_cell_value(self, row: int, col: int) -> str:
        """获取原始单元格值"""
        try:
            cell = self.worksheet.cell(row=row+1, column=col+1)
            return str(cell.value) if cell.value is not None else ""
        except:
            return ""

    @pyqtSlot()
    def save_changes(self):
        """保存更改到worksheet"""
        # 确保在主线程执行保存操作
        if threading.current_thread() is threading.main_thread():
            self._save()
        else:
            QMetaObject.invokeMethod(self, "save_changes", Qt.ConnectionType.QueuedConnection)

    def _save(self):
        """实际执行保存操作的私有方法"""
        with self._save_lock:  # 使用线程锁保护保存操作
            try:
                self._logger.info("保存更改到worksheet")
                for (row, col), value in self._data.items():
                    cell = self.worksheet.cell(row=row+1, column=col+1)
                    if not cell.protection.locked:  # 检查单元格是否只读
                        cell.value = value
                        
            except Exception as e:
                logging.error(f"保存表格数据时发生错误: {e.with_traceback(None)}")
                ErrorHandler.handle_error(e, None, f"保存表格数据时发生错误：{e.with_traceback(None)}")

    def set_cell_color(self, row: int, col: int, color: QColor):
        """设置单元格颜色"""
        self._colors[(row, col)] = color
        index = self.index(row, col)
        self.dataChanged.emit(index, index, [Qt.ItemDataRole.BackgroundRole])
        
    def get_cell_color(self, row: int, col: int) -> Optional[QColor]:
        """获取单元格颜色"""
        return self._colors.get((row, col))
        
    def clear_cell_color(self, row: int, col: int):
        """清除单元格颜色"""
        if (row, col) in self._colors:
            del self._colors[(row, col)]
            index = self.index(row, col)
            self.dataChanged.emit(index, index, [Qt.ItemDataRole.BackgroundRole]) 
        
    def has_changes(self) -> bool:
        """检查是否有未保存的更改"""
        return self._data != self._original_data

    def batch_update_cells(self, column: int, updates: List[dict]):
        """批量更新单元格数据和颜色
        
        Args:
            column: 要更新的列
            updates: 更新数据列表，每项包含 row, value, color
        """
        try:
            # 开始批量更新
            self.beginResetModel()
            
            # 一次性更新所有数据
            for update in updates:
                row = update['row']
                # 更新数据 - 修复这里的错误
                self._data[(row, column)] = str(update['value'])  # 使用元组作为键，并确保值是字符串
                # 更新颜色
                self._colors[(row, column)] = update['color']
                
            # 完成批量更新
            self.endResetModel()
            
            # 发出数据改变信号
            if updates:  # 确保有更新数据
                min_row = min(u['row'] for u in updates)
                max_row = max(u['row'] for u in updates)
                self.dataChanged.emit(
                    self.index(min_row, column),
                    self.index(max_row, column),
                    [Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.BackgroundRole]
                )
            
        except Exception as e:
            logging.error(f"批量更新单元格时发生错误: {str(e)}")
            raise

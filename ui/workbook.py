from typing import Optional, Dict, List
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QTabWidget, 
                            QTableView, QHeaderView, QMessageBox, QProgressDialog)
from PyQt6.QtCore import Qt, QAbstractTableModel
from globals import GlobalState
from utils.error_handler import ErrorHandler
from models.table_model import TableModel
from PyQt6.QtWidgets import QApplication
from openpyxl import Workbook
import logging

class WorkbookWidget(QWidget):
    """工作簿窗口组件，管理Excel工作表的显示"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.tab_widget = QTabWidget()
        self.layout.addWidget(self.tab_widget)
        self._logger = logging.getLogger(__name__)
        
    def load_workbook(self, workbook: Workbook, file_path: str):
        """加载工作簿"""
        try:
            # 检查是否已经加载了相同的工作簿
            state = GlobalState()
            if state.workbook.file_path == file_path:
                return
            
            # 清除现有标签页并释放资源
            self.clear_tabs()
            
            # 更新全局状态
            state.workbook.workbook = workbook
            state.workbook.file_path = file_path
            state.workbook.tab_widget = self.tab_widget
            
            # 使用 QProgressDialog 显示加载进度
            progress = QProgressDialog("正在加载工作表...", None, 0, len(workbook.sheetnames), self)
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            
            # 加载所有工作表
            for i, sheet_name in enumerate(workbook.sheetnames):
                progress.setValue(i)
                progress.setLabelText(f"正在加载工作表: {sheet_name}")
                
                # 创建表格视图
                table_view = QTableView()
                table_view.setModel(TableModel(workbook[sheet_name]))
                
                # 优化表格视图性能
                table_view.setHorizontalScrollMode(QTableView.ScrollMode.ScrollPerPixel)
                table_view.setVerticalScrollMode(QTableView.ScrollMode.ScrollPerPixel)
                table_view.horizontalHeader().setDefaultSectionSize(100)
                table_view.verticalHeader().setDefaultSectionSize(25)
                
                # 添加到标签页
                self.tab_widget.addTab(table_view, sheet_name)
                
                # 处理事件，保持界面响应
                QApplication.processEvents()
                
            progress.setValue(len(workbook.sheetnames))
            
            # 更新工作表名称列表
            state.workbook.sheet_names = workbook.sheetnames
            
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "加载工作簿时发生错误")
        finally:
            progress.close()
            
    def add_sheet_tab(self, sheet, sheet_name: str):
        """
        添加工作表标签页
        
        Args:
            sheet: openpyxl worksheet对象
            sheet_name: 工作表名称
        """
        try:
            # 创建表格视图
            table_view = QTableView()
            
            # 创建数据模型
            model = TableModel(sheet)
            table_view.setModel(model)
            
            # 配置表格视图
            self.setup_table_view(table_view)
            
            # 添加标签页
            self.tab_widget.addTab(table_view, sheet_name)
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, f"添加工作表 {sheet_name} 时发生错误")
            
    def setup_table_view(self, table_view: QTableView):
        """
        配置表格视图
        
        Args:
            table_view: QTableView实例
        """
        # 设置选择模式
        table_view.setSelectionMode(QTableView.SelectionMode.ExtendedSelection)
        table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectItems)
        
        # 设置表头
        horizontal_header = table_view.horizontalHeader()
        horizontal_header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        horizontal_header.setStretchLastSection(True)
        
        vertical_header = table_view.verticalHeader()
        vertical_header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # 允许拖放
        table_view.setDragEnabled(True)
        table_view.setAcceptDrops(True)
        table_view.setDropIndicatorShown(True)
        
        # 其他设置
        # table_view.setAlternatingRowColors(True)
        table_view.setShowGrid(True)
        table_view.setCornerButtonEnabled(True)
        
    def clear_tabs(self):
        """清除所有标签页"""
        while self.tab_widget.count() > 0:
            self.tab_widget.removeTab(0)
            
    def close_tab(self, index: int):
        """
        关闭指定标签页
        
        Args:
            index: 标签页索引
        """
        try:
            # 获取标签页中的表格视图
            table_view = self.tab_widget.widget(index)
            if isinstance(table_view, QTableView):
                # 清理表格视图资源
                model = table_view.model()
                if model:
                    model.deleteLater()
                table_view.deleteLater()
            
            # 移除标签页
            self.tab_widget.removeTab(index)
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, "关闭标签页时发生错误")
            
    def on_tab_changed(self, index: int):
        """
        处理标签页切换事件
        
        Args:
            index: 新的活动标签页索引
        """
        try:
            # 更新当前活动的工作表
            state = GlobalState()
            if index >= 0:
                sheet_name = self.tab_widget.tabText(index)
                if state.workbook.workbook:
                    state.workbook.activate_sheet = state.workbook.workbook[sheet_name]
                    
                # 更新当前活动的表格视图
                current_widget = self.tab_widget.widget(index)
                if isinstance(current_widget, QTableView):
                    self._logger.info("切换到表格视图")
                    state.set_current_table_view(current_widget)
            else:
                state.workbook.activate_sheet = None
                state.set_current_table_view(None)
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "切换标签页时发生错误")
            self._logger.info(f"切换标签页时发生错误：{e}")
            
    def get_current_table_view(self) -> Optional[QTableView]:
        """
        获取当前活动的表格视图
        
        Returns:
            Optional[QTableView]: 当前活动的表格视图，如果没有则返回None
        """
        current_widget = self.tab_widget.currentWidget()
        if isinstance(current_widget, QTableView):
            return current_widget
        return None
        
    def save_workbook(self) -> bool:
        """
        保存工作簿
        
        Returns:
            bool: 是否保存成功
        """
        try:
            state = GlobalState()
            if not state.workbook.workbook or not state.workbook.file_path:
                return False
                
            # 保存所有表格视图的更改到工作簿
            for i in range(self.tab_widget.count()):
                table_view = self.tab_widget.widget(i)
                if isinstance(table_view, QTableView):
                    model = table_view.model()
                    if isinstance(model, TableModel):
                        model.save_changes()
                        
            # 保存工作簿
            state.workbook.workbook.save(state.workbook.file_path)
            return True
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, "保存工作簿时发生错误")
            return False
